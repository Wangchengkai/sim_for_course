# app.py
# -*- coding: utf-8 -*-
"""
急诊仿真网页（Streamlit）
- 固定读取到达率：2025 服务系统问题-问题数据.xlsx（工作表“数据”，第6~12行×第2~25列 -> 7×24）
- 上传 res_doctor.xlsx（标准 .xlsx），切第2~19行×第2~169列 -> (18×168)
- 运行事件驱动仿真（口径A）并输出四项指标
"""

import io
import heapq
from collections import deque
from typing import Tuple, Optional

import numpy as np
import pandas as pd
import streamlit as st
from pathlib import Path

# -------------------------
# 页面设置
# -------------------------
st.set_page_config(page_title="ED 仿真评估", page_icon="⚕️", layout="wide")
st.title("⚕️ 急诊排班仿真评估")
st.caption("只需上传 res_doctor.xlsx；到达率固定。")

# -------------------------
# 路径（以脚本目录为基准）
# -------------------------
BASE_DIR = Path(__file__).resolve().parent
ARRIVAL_DEFAULT_PATH = BASE_DIR / "2025 服务系统问题-问题数据.xlsx"

# -------------------------
# 读取/解析工具
# -------------------------
@st.cache_data(show_spinner=False)
def load_arrival_rates_from_excel(path: Path, sheet_name: str = "数据") -> np.ndarray:
    """读取到达率：7×24"""
    df = pd.read_excel(path, sheet_name=sheet_name, header=None)
    arr = df.iloc[5:12, 1:25].to_numpy(dtype=float)
    if arr.shape != (7, 24):
        raise ValueError(f"到达率矩阵应为 (7,24)，当前为 {arr.shape}。")
    return arr

def _compute_borrow_count(df: pd.DataFrame) -> int:
    """
    借调人数：第13行及以后（0-based>=12），任一工时>0 计1人。
    逐列数值化以稳健处理空/文本等。
    """
    if df.shape[0] <= 12:
        return 0
    borrow_part = df.iloc[12:, 1:]                       # 去掉第一列
    borrow_part = borrow_part.apply(pd.to_numeric, errors="coerce").fillna(0)
    return int((borrow_part.sum(axis=1) > 0).sum())

def _schedule_matrix_from_df(df: pd.DataFrame) -> np.ndarray:
    """切 18×168（第2~19行×第2~169列），并做数值化→四舍五入→int"""
    if df.shape[0] < 19 or df.shape[1] < 169:
        raise ValueError(f"排班表尺寸过小：{df.shape}，至少需 19×169。")
    block = df.iloc[1:19, 1:169]
    block_num = block.apply(pd.to_numeric, errors="coerce").fillna(0)
    sched = block_num.round().astype(int).to_numpy()
    if sched.shape != (18, 168):
        raise ValueError(f"排班矩阵应为 (18,168)，实际为 {sched.shape}。")
    return sched

@st.cache_data(show_spinner=False)
def load_res_doctor_xlsx(content: bytes, filename: str = "res_doctor.xlsx") -> Tuple[np.ndarray, int]:
    """
    稳健读取 res_doctor.xlsx（仅支持标准 .xlsx）：
      - 确认为真正的 .xlsx（zip 头 PK\x03\x04）
      - 用 openpyxl 读取第一张工作表
      - 转成 DataFrame 后切 18×168
      - 统计借调人数
    """
    from openpyxl import load_workbook

    bio = io.BytesIO(content)

    # 1) 必须是真 .xlsx（zip 容器）
    head = bio.read(4)
    bio.seek(0)
    if head != b"PK\x03\x04":
        raise ValueError(
            f"文件 {filename} 不是标准 .xlsx（zip）格式；"
            "请用 Excel 打开后【另存为 → Excel 工作簿(.xlsx)】再上传。"
        )

    # 2) 用 openpyxl 读取
    try:
        wb = load_workbook(bio, data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"openpyxl 无法打开 {filename}：{e}")

    sheet_names = wb.sheetnames or []
    if not sheet_names:
        raise ValueError(f"{filename} 未检测到任何工作表（可能为空白/受保护）。")

    ws = wb[sheet_names[0]]
    values = [[cell.value for cell in row] for row in ws.iter_rows()]
    df = pd.DataFrame(values)

    # 3) 切排班矩阵 + 借调人数
    sched = _schedule_matrix_from_df(df)
    borrow = _compute_borrow_count(df)
    return sched, borrow

# -------------------------
# 仿真核心
# -------------------------
class Simulator:
    def __init__(self, arrival_rates: np.ndarray, service_rate_per_doctor: float = 6.0,
                 n_simulations: int = 1000, seed_base: int = 0):
        self.arrival_rates = arrival_rates
        self.service_rate_per_doctor = float(service_rate_per_doctor)
        self.n_simulations = int(n_simulations)
        self.seed_base = int(seed_base)

    def run_for_schedule(self, doctor_schedule: np.ndarray, borrow_doctor_count: int) -> dict:
        n_doctors, n_hours = doctor_schedule.shape
        T_end = float(n_hours)

        total_arrivals = float(self.arrival_rates.sum())
        hourly_doctors = doctor_schedule.sum(axis=0)                 # (168,)
        total_capacity = float(np.dot(hourly_doctors, np.full(n_hours, self.service_rate_per_doctor)))

        EVENT_ARRIVAL, EVENT_CHECK, EVENT_FINISH = 0, 1, 2

        def on_shift(d, t):
            idx = int(t)
            return (0 <= idx < n_hours) and (doctor_schedule[d, idx] == 1)

        def simulate_once(seed: Optional[int] = None) -> float:
            if seed is not None:
                np.random.seed(seed)

            waiting_queue = deque()
            total_wait_time = 0.0
            doctor_busy_until = np.zeros(n_doctors, dtype=float)
            events = []

            # 生成到达
            for hour in range(n_hours):
                day = hour // 24
                hr_in_day = hour % 24
                lam = float(self.arrival_rates[day % 7, hr_in_day])
                if lam <= 0:
                    continue
                t = 0.0
                while True:
                    inter_arrival = np.random.exponential(1.0 / lam)
                    t += inter_arrival
                    if t >= 1.0:
                        break
                    arrival_time = hour + t
                    service_time = np.random.exponential(1.0 / self.service_rate_per_doctor)
                    heapq.heappush(events, (arrival_time, EVENT_ARRIVAL, "arrival", None, (arrival_time, service_time)))

            # 整点检查
            for hour in range(n_hours):
                heapq.heappush(events, (float(hour), EVENT_CHECK, "check_shift", None, None))

            while events:
                current_time, _, event_type, doctor_id, payload = heapq.heappop(events)

                if event_type == "arrival":
                    arrival_time, service_time = payload
                    waiting_queue.append((arrival_time, service_time))
                    for d in range(n_doctors):
                        if not waiting_queue: break
                        if on_shift(d, current_time) and doctor_busy_until[d] <= current_time:
                            start_time = max(current_time, doctor_busy_until[d])
                            if on_shift(d, start_time):
                                pat_arrival, pat_service = waiting_queue.popleft()
                                finish_time = start_time + pat_service
                                doctor_busy_until[d] = finish_time
                                total_wait_time += (start_time - pat_arrival)
                                heapq.heappush(events, (finish_time, EVENT_FINISH, "finish", d, None))

                elif event_type == "finish":
                    d = doctor_id
                    if d is not None and on_shift(d, current_time) and waiting_queue:
                        start_time = max(current_time, doctor_busy_until[d])
                        if on_shift(d, start_time):
                            pat_arrival, pat_service = waiting_queue.popleft()
                            finish_time = start_time + pat_service
                            doctor_busy_until[d] = finish_time
                            total_wait_time += (start_time - pat_arrival)
                            heapq.heappush(events, (finish_time, EVENT_FINISH, "finish", d, None))

                elif event_type == "check_shift":
                    for d in range(n_doctors):
                        if not waiting_queue: break
                        if on_shift(d, current_time) and doctor_busy_until[d] <= current_time:
                            start_time = max(current_time, doctor_busy_until[d])
                            if on_shift(d, start_time):
                                pat_arrival, pat_service = waiting_queue.popleft()
                                finish_time = start_time + pat_service
                                doctor_busy_until[d] = finish_time
                                total_wait_time += (start_time - pat_arrival)
                                heapq.heappush(events, (finish_time, EVENT_FINISH, "finish", d, None))

            # 口径A：只统计周内等待
            total_wait_time += sum(max(0.0, T_end - arr) for (arr, _) in waiting_queue)
            return total_wait_time

        # 多次仿真
        all_total_wait = np.zeros(self.n_simulations, dtype=float)
        for i in range(self.n_simulations):
            all_total_wait[i] = simulate_once(seed=self.seed_base + i)

        total_doctor_hours = float(doctor_schedule.sum())
        avg_total_wait = float(all_total_wait.mean())
        std_total_wait = float(all_total_wait.std(ddof=1)) if self.n_simulations > 1 else 0.0

        note = ("⚠️ 供给 < 需求：周末会积压，口径A尾巴等待较大。"
                if total_capacity < total_arrivals else
                "✅ 供给 ≥ 需求：尾巴不应过大；若等待仍陡增，请排查数据/事件逻辑。")

        return {
            "avg_total_wait": avg_total_wait,
            "std_total_wait": std_total_wait,
            "total_doctor_hours": total_doctor_hours,
            "borrow_doctor_count": borrow_doctor_count,
            "total_cost": avg_total_wait + total_doctor_hours * 1.3 + borrow_doctor_count * 20.0,
            "note": note
        }

# -------------------------
# 侧边栏参数
# -------------------------
st.sidebar.header("参数设置")
mu    = st.sidebar.number_input("单医生服务率 μ (人/小时)", min_value=0.1, max_value=60.0, value=6.0, step=0.1)
nruns = st.sidebar.number_input("仿真次数", min_value=1, max_value=5000, value=1000, step=100)
seed0 = st.sidebar.number_input("随机种子基数", min_value=0, max_value=10_000_000, value=0, step=1)

# 只需要上传 res_doctor.xlsx
st.subheader("上传医生排班表（res_doctor.xlsx）注意格式一致，以及不要超过18名医生")
upload = st.file_uploader("选择 Excel 文件（必须为 .xlsx 工作簿）", type=["xlsx"], accept_multiple_files=False)

# -------------------------
# 主逻辑
# -------------------------
if upload is not None:
    try:
        # 到达率（固定本地文件）
        if not ARRIVAL_DEFAULT_PATH.exists():
            raise FileNotFoundError(f"未找到到达率文件：{ARRIVAL_DEFAULT_PATH}")
        arrival_rates = load_arrival_rates_from_excel(ARRIVAL_DEFAULT_PATH)

        # 来访者排班
        if upload.size == 0:
            raise ValueError("上传的排班文件为空（0 字节）。")
        sched_user, borrow_user = load_res_doctor_xlsx(upload.getvalue(), filename=upload.name)

        # 运行仿真
        sim = Simulator(arrival_rates, service_rate_per_doctor=mu, n_simulations=nruns, seed_base=seed0)
        with st.spinner("正在运行仿真……"):
            res_user = sim.run_for_schedule(sched_user, borrow_user)

        # 展示结果
        st.subheader("结果（备注：我们不记录168小时后患者的等待，本代码不对排班校验排班可行性）")
        # st.write(res_user["note"])
        st.code(
            f"总等待时间(人·小时) [平均±标准差 over {nruns} runs]: "
            f"{res_user['avg_total_wait']:.2f} ± {res_user['std_total_wait']:.2f}\n"
            f"总工作成本: {res_user['total_doctor_hours']*1.3:.2f}\n"
            f"借调成本: {res_user['borrow_doctor_count']*20.0:.2f}\n"
            f"总成本: {res_user['total_cost']:.2f}",
            language="text"
        )

        # 下载
        out_df = pd.DataFrame([{
            "方案": "当前医生排班",
            "平均总等待(人·小时)": res_user["avg_total_wait"],
            "等待标准差": res_user["std_total_wait"],
            "总工作时长": res_user["total_doctor_hours"],
            "借调人数": res_user["borrow_doctor_count"],
            "总工作成本": res_user["total_doctor_hours"] * 1.3,
            "借调成本": res_user["borrow_doctor_count"] * 20.0,
            "总成本": res_user["total_cost"],
        }])
        st.dataframe(out_df, use_container_width=True)
        csv = out_df.to_csv(index=False).encode("utf-8-sig")
        st.download_button("下载结果 CSV", data=csv, file_name="simulation_results.csv", mime="text/csv")

    except Exception as e:
        st.error("出错了：" + str(e))
        st.info(
            "调试信息：\n"
            f"- 上传文件名：{upload.name if upload else '（未上传）'}\n"
            f"- 固定到达率路径：{ARRIVAL_DEFAULT_PATH}"
        )
else:
    st.info("请上传 res_doctor.xlsx（必须为 .xlsx 工作簿）。")
